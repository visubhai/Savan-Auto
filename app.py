"""
Savan Travels WhatsApp Sender — Flask web application.
Run: python app.py  →  http://localhost:5000
Default login: admin / savan123
"""
import os
import json
import io
import csv
import tempfile
from datetime import datetime
from functools import wraps
from flask import (
    Flask, render_template, request, redirect, url_for, session,
    jsonify, flash, send_file, abort,
)
# `redirect` is already imported above — used by serve_media for R2.

import database as db
from parsers import auto_parse, extract_all_phones, parse_review_export
from meta_api import WhatsAppAPI
import scheduler
import storage
import gmail_fetch

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "savan-travels-localhost-secret-2024")
app.config["MAX_CONTENT_LENGTH"] = 20 * 1024 * 1024  # 20 MB


# ── gzip compression — typical HTML payloads shrink ~70% ────────────────────
@app.after_request
def _gzip_response(resp):
    """Compress textual responses when the client supports gzip."""
    # Skip if already encoded, streaming, or too small to be worth it
    if (
        resp.direct_passthrough
        or resp.headers.get("Content-Encoding")
        or "gzip" not in (request.headers.get("Accept-Encoding") or "")
        or (resp.content_length or 0) < 512
    ):
        return resp
    ct = (resp.content_type or "").split(";")[0].strip().lower()
    if not (ct.startswith("text/") or ct in (
        "application/json", "application/javascript",
        "application/xml", "image/svg+xml",
    )):
        return resp
    try:
        import gzip
        data = resp.get_data()
        gz = gzip.compress(data, compresslevel=6)
        if len(gz) >= len(data):
            return resp  # not worth it
        resp.set_data(gz)
        resp.headers["Content-Encoding"] = "gzip"
        resp.headers["Content-Length"] = str(len(gz))
        vary = resp.headers.get("Vary")
        resp.headers["Vary"] = (vary + ", Accept-Encoding") if vary else "Accept-Encoding"
    except Exception:
        pass
    return resp


# ── light caching for static assets ────────────────────────────────────────
@app.after_request
def _cache_static(resp):
    if request.path.startswith("/static/"):
        resp.headers["Cache-Control"] = "public, max-age=86400"
    return resp

# FIX 1: Use filesystem session storage for large CSVs instead of cookie
# Store pending passengers in a temp file, only store path in session
TEMP_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(TEMP_DIR, exist_ok=True)

# Incoming WhatsApp media (images, audio, video, documents) is saved here.
# Note: on Render free tier this disk is ephemeral and wiped on redeploy.
MEDIA_DIR = os.path.join(TEMP_DIR, "media")
os.makedirs(MEDIA_DIR, exist_ok=True)


# Maps MIME types we expect from WhatsApp to safe file extensions
_MIME_EXT = {
    "image/jpeg": ".jpg", "image/jpg": ".jpg", "image/png": ".png",
    "image/webp": ".webp", "image/gif": ".gif",
    "audio/ogg": ".ogg", "audio/mpeg": ".mp3", "audio/mp4": ".m4a",
    "audio/amr": ".amr", "audio/aac": ".aac",
    "video/mp4": ".mp4", "video/3gpp": ".3gp",
    "application/pdf": ".pdf",
    "application/msword": ".doc",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
    "application/vnd.ms-excel": ".xls",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    "text/plain": ".txt",
}


def render_template_body(body, name=None, route=None, platform=None, extra=None, template_name=None):
    """Substitute {{1}}, {{2}}, {{3}}, … in a WhatsApp template body.

    Mapping varies by template type — must mirror scheduler.send_batch:
      • Follow-up templates (name contains "follow"): {{1}}=platform,
        {{2}}=name, {{3}}=route
      • All other templates:                          {{1}}=name,
        {{2}}=route, {{3}}=platform

    `extra` is a dict of {position_str: value} for any higher-indexed
    variables (e.g. fixed campaign values) — used opportunistically;
    unknown placeholders are left as-is.
    """
    if not body:
        return ""
    import re
    is_followup = bool(template_name and "follow" in template_name.lower())
    if is_followup:
        subs = {
            "1": (platform or "the platform"),
            "2": (name or "Customer"),
            "3": (route or "your journey"),
        }
    else:
        subs = {
            "1": (name or "Customer"),
            "2": (route or "your journey"),
            "3": (platform or "the platform"),
        }
    if extra:
        for k, v in extra.items():
            if v:
                subs[str(k)] = str(v)
    def _sub(m):
        return subs.get(m.group(1), m.group(0))
    return re.sub(r"\{\{(\d+)\}\}", _sub, body)


def _ext_for_mime(mime):
    """Best-effort file extension for a MIME type from WhatsApp."""
    if not mime:
        return ".bin"
    base = mime.split(";")[0].strip().lower()
    if base in _MIME_EXT:
        return _MIME_EXT[base]
    import mimetypes
    return mimetypes.guess_extension(base) or ".bin"

# Custom Jinja filters
@app.template_filter("from_json")
def from_json_filter(s):
    try:
        return json.loads(s)
    except Exception:
        return []

# Initialize DB on startup
db.init_db()

# Start background scheduler
scheduler.start_scheduler()


# ---------------- Session helpers (FIX 1: file-based pending store) ----------------
def save_pending(data):
    """Save large pending data to temp file, return file key."""
    key = f"pending_{session.get('user_id', 0)}_{db.now_ist().strftime('%Y%m%d%H%M%S%f')}.json"
    path = os.path.join(TEMP_DIR, key)
    with open(path, "w") as f:
        json.dump(data, f)
    # Clean old pending files for this user
    prefix = f"pending_{session.get('user_id', 0)}_"
    for fn in os.listdir(TEMP_DIR):
        if fn.startswith(prefix) and fn != key:
            try:
                os.remove(os.path.join(TEMP_DIR, fn))
            except Exception:
                pass
    return key


def load_pending(key):
    """Load pending data from temp file."""
    if not key:
        return None
    path = os.path.join(TEMP_DIR, key)
    if not os.path.exists(path):
        return None
    try:
        with open(path) as f:
            return json.load(f)
    except Exception:
        return None


def clear_pending(key):
    if not key:
        return
    path = os.path.join(TEMP_DIR, key)
    try:
        os.remove(path)
    except Exception:
        pass


def _safe_int(val, default=0, minimum=None):
    """Parse an int from untrusted input without crashing."""
    try:
        n = int(str(val).strip())
    except (TypeError, ValueError):
        return default
    if minimum is not None and n < minimum:
        return minimum
    return n


def _parse_var_overrides(raw):
    """Parse the var_overrides JSON from a form field into a clean dict.

    Returns {var_index_str: fixed_value} keeping only non-empty values.
    e.g. '{"3":"RedBus","4":"Diwali Offer"}' -> {'3':'RedBus','4':'Diwali Offer'}
    """
    raw = (raw or "").strip()
    if not raw:
        return {}
    try:
        data = json.loads(raw)
    except Exception:
        return {}
    if not isinstance(data, dict):
        return {}
    clean = {}
    for k, v in data.items():
        val = str(v).strip() if v is not None else ""
        if val:
            clean[str(k)] = val
    return clean


# ---------------- Auth decorators ----------------
def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login", next=request.path))
        return f(*args, **kwargs)
    return wrapper


def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if "user_id" not in session:
            return redirect(url_for("login", next=request.path))
        if session.get("role") != "admin":
            flash("Admin access required", "error")
            return redirect(url_for("dashboard"))
        return f(*args, **kwargs)
    return wrapper


@app.context_processor
def inject_user():
    return {
        "current_user": {
            "id": session.get("user_id"),
            "username": session.get("username"),
            "display_name": session.get("display_name"),
            "role": session.get("role"),
        } if "user_id" in session else None,
        "business_name": db.get_setting("business_name", "Savan Travels"),
    }


# ---------------- Auth routes ----------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = db.get_user_by_username(username)
        if user and db.verify_password(password, user["password_hash"]):
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["display_name"] = user["display_name"]
            session["role"] = user["role"]
            next_url = request.args.get("next") or url_for("dashboard")
            return redirect(next_url)
        flash("Invalid username or password", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    # Clean up pending file
    clear_pending(session.get("pending_key"))
    session.clear()
    return redirect(url_for("login"))


@app.route("/")
def index():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))


# ---------------- Dashboard ----------------
@app.route("/dashboard")
@login_required
def dashboard():
    today = db.get_today_stats()
    month = db.get_month_stats()
    chart_data = db.get_chart_data(7)
    recent = db.get_recent_sends(10)
    top_routes = db.get_top_routes(5)

    from datetime import timedelta
    today_date = db.now_ist().date()
    chart_map = {row["day"]: row for row in chart_data}
    days = []
    for i in range(6, -1, -1):
        d = today_date - timedelta(days=i)
        d_str = d.isoformat()
        row = chart_map.get(d_str, {})
        days.append({
            "day": d.strftime("%d %b"),
            "sent": row.get("sent", 0),
            "failed": row.get("failed", 0),
        })

    api = WhatsAppAPI()
    api_configured = api.is_configured()
    batches = db.list_batches(5)

    return render_template("dashboard.html",
                           today=today, month=month, chart=days,
                           recent=recent, top_routes=top_routes,
                           api_configured=api_configured,
                           batches=batches)


# ---------------- Send messages ----------------
ALLOWED_EXTENSIONS = {".csv", ".txt", ".tsv"}

@app.route("/send", methods=["GET", "POST"])
@login_required
def send():
    templates = [dict(t) for t in db.get_templates()]
    default_tpl = db.get_default_template()
    default_tpl = dict(default_tpl) if default_tpl else None

    if request.method == "POST":
        files = request.files.getlist("csv_file")
        files = [f for f in files if f.filename]

        if not files:
            flash("No file uploaded", "error")
            return redirect(url_for("send"))

        all_passengers = []
        seen_phones_global = set()
        all_stats = {"total_rows": 0, "cancelled": 0, "duplicates": 0, "invalid": 0}
        filenames = []

        for f in files:
            ext = os.path.splitext(f.filename)[1].lower()
            if ext not in ALLOWED_EXTENSIONS:
                flash(f"'{f.filename}' skipped — only .csv/.txt/.tsv files are allowed.", "error")
                continue
            try:
                content = f.read()
                parsed = auto_parse(content, filename=f.filename)
            except Exception as e:
                flash(f"Failed to parse '{f.filename}': {e}", "error")
                continue

            filenames.append(f.filename)
            all_stats["total_rows"] += parsed["total_rows_seen"]
            all_stats["cancelled"] += parsed["cancelled_count"]
            all_stats["duplicates"] += parsed["duplicates_removed"]
            all_stats["invalid"] += parsed["invalid_phones"]

            for p in parsed["passengers"]:
                if p["phone"] not in seen_phones_global:
                    seen_phones_global.add(p["phone"])
                    all_passengers.append(p)
                else:
                    all_stats["duplicates"] += 1

        if not filenames:
            flash("No valid CSV files were uploaded.", "error")
            return redirect(url_for("send"))

        if not all_passengers and all_stats["total_rows"] == 0:
            flash("CSV appears empty or has no recognizable data. Check the file has phone numbers.", "error")
            return redirect(url_for("send"))

        # Filter opted-out customers
        opted_out_rows = db.search_customers(opted_out=True, limit=50000)
        opted_out_phones = {r["phone"] for r in opted_out_rows}
        original_count = len(all_passengers)
        all_passengers = [p for p in all_passengers if p["phone"] not in opted_out_phones]
        opt_out_filtered = original_count - len(all_passengers)

        if len(all_passengers) > 2000:
            delay = float(db.get_setting("delay_between_messages", "1.5"))
            hours = round(len(all_passengers) * delay / 3600, 1)
            flash(f"⚠ Large batch: {len(all_passengers)} messages will take ~{hours} hours to send. Consider splitting.", "info")

        batch_label = " + ".join(filenames) if len(filenames) > 1 else filenames[0]

        pending_data = {
            "passengers": all_passengers,
            "stats": {
                "total_rows": all_stats["total_rows"],
                "cancelled": all_stats["cancelled"],
                "duplicates": all_stats["duplicates"],
                "invalid": all_stats["invalid"],
                "opted_out": opt_out_filtered,
                "ready": len(all_passengers),
            },
            "filename": batch_label,
        }
        key = save_pending(pending_data)
        session["pending_key"] = key
        return redirect(url_for("send_preview"))

    return render_template("send.html", templates=templates, default_tpl=default_tpl)


@app.route("/send/preview")
@login_required
def send_preview():
    key = session.get("pending_key")
    pending = load_pending(key)
    if not pending:
        flash("Session expired or no file uploaded. Please upload again.", "error")
        return redirect(url_for("send"))

    templates = [dict(t) for t in db.get_templates()]
    # If the pending batch hinted at a specific template (e.g. Review
    # Follow-up wants "review_followup"), pre-select that one. Otherwise
    # fall back to the DB-wide default template.
    default_tpl = None
    hinted_name = pending.get("default_template")
    if hinted_name:
        default_tpl = next(
            (t for t in templates if t.get("name") == hinted_name),
            None,
        )
    if not default_tpl:
        row = db.get_default_template()
        default_tpl = dict(row) if row else None

    cost_per = float(db.get_setting("cost_per_message", "0.12"))
    estimated_cost = round(len(pending["passengers"]) * cost_per, 2)
    delay = float(db.get_setting("delay_between_messages", "1.5"))
    estimated_mins = round(len(pending["passengers"]) * delay / 60, 1)

    # FIX 7: flag large batches for extra confirmation
    needs_confirm = len(pending["passengers"]) > 200

    return render_template("send_preview.html",
                           pending=pending, templates=templates,
                           default_tpl=default_tpl,
                           estimated_cost=estimated_cost,
                           estimated_mins=estimated_mins,
                           needs_confirm=needs_confirm)


@app.route("/send/start", methods=["POST"])
@login_required
def send_start():
    key = session.get("pending_key")
    pending = load_pending(key)
    if not pending:
        flash("Session expired. Please upload CSV again.", "error")
        return redirect(url_for("send"))

    template_name = request.form.get("template_name") or request.form.get("hiddenTpl")
    if not template_name:
        flash("Please select a template", "error")
        return redirect(url_for("send_preview"))

    # FIX 4: Always read fresh token from DB
    api = WhatsAppAPI()
    if not api.is_configured():
        flash("WhatsApp API token not configured. Go to Settings first.", "error")
        return redirect(url_for("settings"))

    # Use edited passenger list from table if provided
    passengers_json = request.form.get("passengers_json", "").strip()
    if passengers_json:
        try:
            passengers = json.loads(passengers_json)
            # Validate each entry has a phone
            passengers = [p for p in passengers if p.get("phone")]
        except Exception:
            passengers = pending["passengers"]
    else:
        passengers = pending["passengers"]

    if not passengers:
        flash("No passengers to send to. Please check your selection.", "error")
        return redirect(url_for("send_preview"))

    # FIX 7: extra confirmation for large batches
    if len(passengers) > 200:
        confirmed = request.form.get("confirmed") == "yes"
        if not confirmed:
            flash("Please confirm the large batch send.", "error")
            return redirect(url_for("send_preview"))

    # Optional fixed variable values for this batch (for template vars not in CSV)
    var_overrides = _parse_var_overrides(request.form.get("var_overrides", ""))

    batch_name = pending.get("filename", "Manual upload")
    batch_id = scheduler.start_send_thread(
        passengers, template_name, batch_name, session["user_id"],
        var_overrides=var_overrides,
    )
    clear_pending(key)
    session.pop("pending_key", None)
    return redirect(url_for("send_progress", batch_id=batch_id))


@app.route("/send/schedule", methods=["POST"])
@login_required
def send_schedule():
    key = session.get("pending_key")
    pending = load_pending(key)
    if not pending:
        flash("Session expired. Please upload CSV again.", "error")
        return redirect(url_for("send"))

    template_name = request.form.get("template_name")
    scheduled_for = request.form.get("scheduled_for")
    if not template_name or not scheduled_for:
        flash("Template and schedule time are required", "error")
        return redirect(url_for("send_preview"))

    try:
        dt = datetime.fromisoformat(scheduled_for)
        if dt <= db.now_ist():
            flash("Scheduled time must be in the future", "error")
            return redirect(url_for("send_preview"))
        scheduled_str = dt.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        flash("Invalid schedule time format", "error")
        return redirect(url_for("send_preview"))

    # Use edited passenger list if provided (same as send_start does)
    passengers_json = request.form.get("passengers_json", "").strip()
    if passengers_json:
        try:
            passengers = [p for p in json.loads(passengers_json) if p.get("phone")]
        except Exception:
            passengers = pending["passengers"]
    else:
        passengers = pending["passengers"]

    # Optional fixed variable values for this batch (for template vars not in CSV)
    var_overrides = _parse_var_overrides(request.form.get("var_overrides", ""))

    job_id = db.create_scheduled_job(
        name=pending.get("filename", "Scheduled batch"),
        csv_data=json.dumps(passengers),
        template_name=template_name,
        scheduled_for=scheduled_str,
        created_by=session["user_id"],
        var_overrides=json.dumps(var_overrides) if var_overrides else "",
    )
    clear_pending(key)
    session.pop("pending_key", None)
    flash(f"✓ Scheduled {len(passengers)} messages for {scheduled_str}", "success")
    return redirect(url_for("scheduled"))


@app.route("/send/progress/<int:batch_id>")
@login_required
def send_progress(batch_id):
    batch = db.get_batch(batch_id)
    if not batch:
        flash("Batch not found", "error")
        return redirect(url_for("dashboard"))
    return render_template("send_progress.html", batch=dict(batch))


@app.route("/send/status/<int:batch_id>")
@login_required
def send_status(batch_id):
    batch = db.get_batch(batch_id)
    if not batch:
        return jsonify({"error": "not found"}), 404
    live = scheduler.RUNNING_BATCHES.get(batch_id, {})

    # FIX 3: If batch is completed in DB but not in RUNNING_BATCHES (server restart),
    # return completed status from DB
    status = batch["status"]
    if status == "completed":
        return jsonify({
            "id": batch["id"],
            "total": batch["total_count"],
            "sent": batch["sent_count"],
            "failed": batch["failed_count"],
            "status": "completed",
            "current_name": "",
            "completed_at": batch["completed_at"],
        })

    return jsonify({
        "id": batch["id"],
        "total": batch["total_count"],
        "sent": batch["sent_count"],
        "failed": batch["failed_count"],
        "status": status,
        "current_name": live.get("current_name", ""),
        "completed_at": batch["completed_at"],
    })


@app.route("/send/retry/<int:batch_id>", methods=["POST"])
@login_required
def send_retry(batch_id):
    failed = db.get_failed_messages(batch_id)
    if not failed:
        flash("No failed messages to retry", "info")
        return redirect(url_for("send_progress", batch_id=batch_id))

    passengers = [{
        "phone": m["customer_phone"],
        "name": m["customer_name"] or "Customer",
        "route": m["route"] or "",
        "platform": m["platform"] or "",
    } for m in failed]

    template_name = failed[0]["template_name"] or db.get_setting("default_template", "journey_reminder")
    new_batch_id = scheduler.start_send_thread(
        passengers, template_name, f"Retry of #{batch_id}", session["user_id"],
    )
    return redirect(url_for("send_progress", batch_id=new_batch_id))


# ---------------- History ----------------
@app.route("/history")
@login_required
def history():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip() or None
    days_str = request.args.get("days", "").strip()
    days = int(days_str) if days_str.isdigit() else None
    page = _safe_int(request.args.get("page", "1"), default=1, minimum=1)
    per_page = 50
    offset = (page - 1) * per_page

    messages = db.search_messages(q, status, days, per_page, offset)
    total = db.count_messages(q, status, days)
    total_pages = max(1, (total + per_page - 1) // per_page)

    return render_template("history.html",
                           messages=messages, q=q, status=status, days=days,
                           page=page, total=total, total_pages=total_pages)


@app.route("/history/export")
@login_required
def history_export():
    q = request.args.get("q", "").strip()
    status = request.args.get("status", "").strip() or None
    days_str = request.args.get("days", "").strip()
    days = int(days_str) if days_str.isdigit() else None
    messages = db.search_messages(q, status, days, limit=200000)

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["sent_at", "phone", "name", "route", "platform",
                     "template", "status", "error", "wa_message_id"])
    for m in messages:
        writer.writerow([
            m["sent_at"], m["customer_phone"], m["customer_name"],
            m["route"], m["platform"], m["template_name"],
            m["status"], m["error_message"] or "", m["wa_message_id"] or "",
        ])
    output.seek(0)
    filename = f"history_{db.now_ist().strftime('%Y%m%d_%H%M%S')}.csv"
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=filename,
    )


# ---------------- Customers ----------------
@app.route("/customers")
@login_required
def customers():
    q = request.args.get("q", "").strip()
    filter_opt = request.args.get("filter", "").strip()
    opted_out = True if filter_opt == "opted_out" else (False if filter_opt == "active" else None)
    page = _safe_int(request.args.get("page", "1"), default=1, minimum=1)
    per_page = 50
    offset = (page - 1) * per_page

    rows = db.search_customers(q, opted_out, per_page, offset)
    total = db.count_customers(q, opted_out)
    total_pages = max(1, (total + per_page - 1) // per_page)

    return render_template("customers.html",
                           customers=rows, q=q, filter_opt=filter_opt,
                           page=page, total=total, total_pages=total_pages)


@app.route("/customers/toggle/<phone>", methods=["POST"])
@login_required
def customer_toggle(phone):
    opted_out = request.form.get("opted_out") == "1"
    db.toggle_opt_out(phone, opted_out)
    action = "opted out" if opted_out else "restored"
    flash(f"Customer {phone} {action}", "success")
    return redirect(request.referrer or url_for("customers"))


# ---------------- Templates ----------------
@app.route("/templates")
@login_required
def templates_page():
    templates = db.get_templates()
    api = WhatsAppAPI()
    return render_template("templates.html",
                           templates=templates,
                           api_configured=api.is_configured())


@app.route("/templates/sync", methods=["POST"])
@login_required
def templates_sync():
    api = WhatsAppAPI()
    if not api.is_configured():
        flash("WhatsApp API not configured. Add token in Settings first.", "error")
        return redirect(url_for("templates_page"))
    success, msg = api.fetch_templates()
    flash(msg, "success" if success else "error")
    return redirect(url_for("templates_page"))


@app.route("/templates/set-default/<name>", methods=["POST"])
@login_required
def templates_set_default(name):
    db.set_default_template(name)
    flash(f"✓ '{name}' set as default template", "success")
    return redirect(url_for("templates_page"))


@app.route("/templates/<name>/upload-image", methods=["POST"])
@login_required
def templates_upload_image(name):
    """Upload a custom header image for an image-header template.

    The template structure stays approved — Meta only locks the header
    FORMAT (=IMAGE), the actual media is provided at send time. So we
    can swap the image any time without re-submitting the template.

    Flow:
      1. Read uploaded file, validate size/MIME
      2. Upload to Meta Media API → fresh media_id (lasts ~30 days)
      3. Store the media_id on the template (header_media_id)
      4. Mark the source as custom so the UI shows it differently
    """
    tpl = db.get_template_by_name(name)
    if not tpl:
        return jsonify({"ok": False, "error": "template not found"}), 404
    tpl = dict(tpl)
    if (tpl.get("header_type") or "").upper() != "IMAGE":
        return jsonify({"ok": False, "error": "template has no image header"}), 400

    files = [f for f in request.files.getlist("image") if f.filename]
    if not files:
        return jsonify({"ok": False, "error": "no file provided"}), 400
    f = files[0]
    content = f.read()
    if not content:
        return jsonify({"ok": False, "error": "empty file"}), 400
    # Meta accepts images up to 5 MB
    if len(content) > 5 * 1024 * 1024:
        return jsonify({"ok": False, "error": "image too large (max 5 MB)"}), 400
    mime = (f.mimetype or "").lower() or "image/jpeg"
    if not mime.startswith("image/"):
        return jsonify({"ok": False, "error": "file must be an image"}), 400

    api = WhatsAppAPI()
    if not api.is_configured():
        return jsonify({"ok": False, "error": "WhatsApp API not configured"}), 400

    media_id = api.upload_media(content, mime, filename=name)
    if not media_id:
        return jsonify({"ok": False, "error": "Meta refused the upload (check size/format)"}), 502

    # Persist new media_id + mark as custom
    db.upsert_template(
        tpl["name"], tpl["language"], tpl["category"], tpl.get("body", ""),
        tpl.get("variable_count", 0), tpl.get("status", "approved"),
        header_type=tpl.get("header_type"),
        header_example=tpl.get("header_example"),
        header_media_id=media_id,
        buttons=tpl.get("buttons"),
    )
    # Track this as a manual upload so sync doesn't overwrite it later
    db.update_template_meta(
        tpl["name"],
        header_image_is_custom=True,
        header_image_size=len(content),
        header_image_mime=mime,
        header_image_uploaded_at=db.now_ist().strftime("%Y-%m-%d %H:%M:%S"),
    )

    return jsonify({"ok": True, "media_id": media_id, "size": len(content)})


# ---------------- Scheduled jobs ----------------
@app.route("/scheduled")
@login_required
def scheduled():
    jobs = db.list_scheduled_jobs()
    return render_template("scheduled.html", jobs=jobs)


@app.route("/scheduled/cancel/<int:job_id>", methods=["POST"])
@login_required
def scheduled_cancel(job_id):
    db.delete_scheduled_job(job_id)
    flash(f"Scheduled job #{job_id} cancelled", "success")
    return redirect(url_for("scheduled"))


# ---------------- Settings ----------------
@app.route("/settings", methods=["GET", "POST"])
@login_required
def settings():
    if request.method == "POST":
        keys = [
            "access_token", "phone_number_id", "waba_id", "api_version",
            "delay_between_messages", "max_retries", "cost_per_message",
            "wapsolution_monthly_cost", "business_name",
            # WhatsApp tier override (Meta API field sometimes lags the UI)
            "whatsapp_tier_limit_override",
            # Gmail auto-fetch (Review Follow-up)
            "gmail_address", "gmail_app_password",
            "gmail_sender_filter", "gmail_subject_filter",
        ]
        for k in keys:
            v = request.form.get(k)
            if v is not None:
                # Don't wipe stored secrets when the field is submitted blank
                if k in ("access_token", "gmail_app_password") and not v.strip():
                    continue
                db.set_setting(k, v.strip())
        flash("✓ Settings saved successfully", "success")
        return redirect(url_for("settings"))

    all_settings = db.get_all_settings()
    return render_template("settings.html", settings=all_settings)


@app.route("/settings/test", methods=["POST"])
@login_required
def settings_test():
    # FIX 4: Always create fresh API instance to pick up latest token
    api = WhatsAppAPI()
    success, msg = api.test_connection()
    return jsonify({"success": success, "message": msg})


# ---------------- Change password ----------------
@app.route("/settings/change-password", methods=["POST"])
@login_required
def change_password():
    current = request.form.get("current_password", "")
    new_pw = request.form.get("new_password", "")
    confirm = request.form.get("confirm_password", "")

    user = db.get_user_by_id(session["user_id"])
    if not db.verify_password(current, user["password_hash"]):
        flash("Current password is incorrect", "error")
        return redirect(url_for("settings"))
    if len(new_pw) < 6:
        flash("New password must be at least 6 characters", "error")
        return redirect(url_for("settings"))
    if new_pw != confirm:
        flash("Passwords don't match", "error")
        return redirect(url_for("settings"))

    db.change_user_password(session["user_id"], db.hash_password(new_pw))
    flash("✓ Password changed successfully", "success")
    return redirect(url_for("settings"))


# ---------------- Auto-replies (customer-facing welcome menu + FAQ) ----------------

@app.route("/auto-replies")
@login_required
def auto_replies():
    rows = db.list_auto_replies()
    enabled       = db.get_setting("auto_reply_enabled", "0") == "1"
    welcome_body  = db.get_setting(
        "welcome_menu_body",
        "Welcome to Savan Travels 🚌\nHow can we help you today?",
    )
    welcome_footer    = db.get_setting("welcome_menu_footer", "")
    list_button_label = db.get_setting("welcome_list_button_label", "Choose an option")
    return render_template(
        "auto_replies.html",
        replies=rows,
        enabled=enabled,
        welcome_body=welcome_body,
        welcome_footer=welcome_footer,
        list_button_label=list_button_label,
    )


@app.route("/auto-replies/save-settings", methods=["POST"])
@login_required
def auto_replies_save_settings():
    db.set_setting("auto_reply_enabled",
                   "1" if request.form.get("enabled") == "on" else "0")
    db.set_setting("welcome_menu_body",        request.form.get("welcome_body", "").strip())
    db.set_setting("welcome_menu_footer",      request.form.get("welcome_footer", "").strip())
    # Only used when the menu has 4+ items (renders as a list message).
    db.set_setting("welcome_list_button_label",
                   (request.form.get("list_button_label", "").strip() or "Choose an option"))
    flash("✓ Auto-reply settings updated", "success")
    return redirect(url_for("auto_replies"))


def _parse_reply_form(form):
    """Pull the fields shared by create + edit auto-reply routes."""
    label    = form.get("trigger_label", "").strip()
    keywords = form.get("keywords", "").strip()
    text     = form.get("response_text", "").strip()
    desc     = form.get("description", "").strip()
    in_menu  = form.get("show_in_menu") == "on"
    try:
        order_int = int(form.get("menu_order", "0").strip() or "0")
    except ValueError:
        order_int = 0
    return label, keywords, text, desc, in_menu, order_int


def _validate_reply_fields(label, text, desc):
    """Return an error string if invalid, else None."""
    if not label or not text:
        return "Trigger label and response text are required"
    # 20 chars covers WhatsApp's button-title limit; 24 covers list-row titles.
    # We cap at the tighter button limit so the same label works in both modes.
    if len(label) > 20:
        return "Trigger label must be 20 characters or fewer (WhatsApp button limit)"
    if desc and len(desc) > 72:
        return "Description must be 72 characters or fewer (WhatsApp list row limit)"
    return None


@app.route("/auto-replies/create", methods=["POST"])
@login_required
def auto_replies_create():
    label, keywords, text, desc, in_menu, order_int = _parse_reply_form(request.form)
    err = _validate_reply_fields(label, text, desc)
    if err:
        flash(err, "error")
        return redirect(url_for("auto_replies"))
    db.create_auto_reply(label, keywords, text, in_menu, order_int, description=desc)
    flash(f"✓ Auto-reply '{label}' added", "success")
    return redirect(url_for("auto_replies"))


@app.route("/auto-replies/<int:reply_id>/edit", methods=["POST"])
@login_required
def auto_replies_edit(reply_id):
    label, keywords, text, desc, in_menu, order_int = _parse_reply_form(request.form)
    err = _validate_reply_fields(label, text, desc)
    if err:
        flash(err, "error")
        return redirect(url_for("auto_replies"))
    db.update_auto_reply(reply_id, label, keywords, text, in_menu, order_int, description=desc)
    flash("✓ Auto-reply updated", "success")
    return redirect(url_for("auto_replies"))


@app.route("/auto-replies/<int:reply_id>/toggle", methods=["POST"])
@login_required
def auto_replies_toggle(reply_id):
    row = db.get_auto_reply(reply_id)
    if not row:
        return jsonify({"ok": False, "error": "not found"}), 404
    db.set_auto_reply_enabled(reply_id, not row["enabled"])
    return jsonify({"ok": True, "enabled": not row["enabled"]})


@app.route("/auto-replies/<int:reply_id>/delete", methods=["POST"])
@login_required
def auto_replies_delete(reply_id):
    db.delete_auto_reply(reply_id)
    flash("Auto-reply deleted", "success")
    return redirect(url_for("auto_replies"))


# ---------------- Team (users) ----------------
@app.route("/users")
@admin_required
def users():
    rows = db.list_users()
    return render_template("users.html", users=rows)


@app.route("/users/create", methods=["POST"])
@admin_required
def users_create():
    username = request.form.get("username", "").strip().lower()
    password = request.form.get("password", "")
    display_name = request.form.get("display_name", "").strip()
    role = request.form.get("role", "member")

    if not username or not password or not display_name:
        flash("All fields are required", "error")
        return redirect(url_for("users"))
    if len(password) < 6:
        flash("Password must be at least 6 characters", "error")
        return redirect(url_for("users"))
    try:
        db.create_user(username, password, display_name, role)
        flash(f"✓ User '{username}' created. They can login at http://localhost:5000", "success")
    except Exception as e:
        if "UNIQUE" in str(e):
            flash(f"Username '{username}' already exists", "error")
        else:
            flash(f"Error creating user: {e}", "error")
    return redirect(url_for("users"))


@app.route("/users/delete/<int:user_id>", methods=["POST"])
@admin_required
def users_delete(user_id):
    if user_id == session["user_id"]:
        flash("Cannot delete your own account", "error")
        return redirect(url_for("users"))
    db.delete_user(user_id)
    flash("User deleted", "success")
    return redirect(url_for("users"))


# ---------------- Batch history ----------------
@app.route("/batches")
@login_required
def batches():
    all_batches = db.list_batches(100)
    return render_template("batches.html", batches=all_batches)


# ---------------- Bulk Campaigns ----------------

@app.route("/campaigns")
@login_required
def campaigns():
    camps = db.list_campaigns()
    templates = db.get_templates()
    # Tag each campaign with its template's header_type so the UI can
    # conditionally show the per-campaign banner upload control.
    tpl_header = {t["name"]: (t.get("header_type") or "") for t in templates}
    for c in camps:
        c["template_header_type"] = tpl_header.get(c.get("template_name"), "")
    return render_template("campaigns.html", campaigns=camps, templates=templates)


def _collect_campaign_variables(form):
    """Pull var_1, var_2, … var_N from a form, in order, with no upper cap.
    Preserves position for empty values so {{N}} placeholders stay aligned
    with what the operator typed."""
    pairs = []
    for key in form.keys():
        if key.startswith("var_"):
            try:
                idx = int(key[4:])
            except ValueError:
                continue
            pairs.append((idx, form.get(key, "").strip()))
    pairs.sort(key=lambda kv: kv[0])
    return [v for _, v in pairs]


@app.route("/campaigns/create", methods=["POST"])
@login_required
def campaigns_create():
    name          = request.form.get("name", "").strip()
    template_name = request.form.get("template_name", "").strip()
    variables = _collect_campaign_variables(request.form)
    if not name or not template_name:
        flash("Campaign name and template are required", "error")
        return redirect(url_for("campaigns"))
    db.create_campaign(name, template_name, variables)
    flash(f"✓ Campaign '{name}' saved", "success")
    return redirect(url_for("campaigns"))


@app.route("/campaigns/<int:campaign_id>/edit", methods=["POST"])
@login_required
def campaigns_edit(campaign_id):
    name          = request.form.get("name", "").strip()
    template_name = request.form.get("template_name", "").strip()
    variables = _collect_campaign_variables(request.form)
    if not name or not template_name:
        flash("Name and template are required", "error")
    else:
        db.update_campaign(campaign_id, name, template_name, variables)
        flash("✓ Campaign updated", "success")
    return redirect(url_for("campaigns"))


@app.route("/campaigns/<int:campaign_id>/delete", methods=["POST"])
@login_required
def campaigns_delete(campaign_id):
    db.delete_campaign(campaign_id)
    flash("Campaign deleted", "success")
    return redirect(url_for("campaigns"))


@app.route("/campaigns/<int:campaign_id>/send", methods=["GET", "POST"])
@login_required
def campaigns_send(campaign_id):
    camp = db.get_campaign(campaign_id)
    if not camp:
        flash("Campaign not found", "error")
        return redirect(url_for("campaigns"))
    template = db.get_template_by_name(camp["template_name"])

    if request.method == "POST":
        # Read phone list submitted from the editable textarea (as JSON)
        phones_json = request.form.get("phones_json", "").strip()
        try:
            all_phones = json.loads(phones_json) if phones_json else []
        except Exception:
            all_phones = []

        # Clean & deduplicate
        from parsers import clean_phone as _cp
        seen = set()
        cleaned = []
        for p in all_phones:
            c = _cp(str(p).strip())
            if c and c not in seen:
                seen.add(c)
                cleaned.append(c)
        all_phones = cleaned

        if not all_phones:
            flash("No valid phone numbers to send to. Add numbers on the right panel.", "error")
            return render_template("campaign_send.html", campaign=camp, template=template)

        # Filter opted-out
        opted_out_phones = {r["phone"] for r in db.search_customers(opted_out=True, limit=50000)}
        all_phones = [p for p in all_phones if p not in opted_out_phones]

        if not all_phones:
            flash("All numbers are opted out.", "error")
            return render_template("campaign_send.html", campaign=camp, template=template)

        passengers = [{"phone": p, "name": "Customer", "route": "", "platform": ""} for p in all_phones]
        batch_name = f"[Campaign] {camp['name']}"
        batch_id = scheduler.start_send_thread(
            passengers, camp["template_name"], batch_name,
            session["user_id"], fixed_params=camp["variables"],
            header_media_id_override=camp.get("header_media_id"),
        )
        db.update_campaign_last_used(campaign_id)
        return redirect(url_for("send_progress", batch_id=batch_id))

    return render_template("campaign_send.html", campaign=camp, template=template)


@app.route("/campaigns/<int:campaign_id>/upload-image", methods=["POST"])
@login_required
def campaigns_upload_image(campaign_id):
    """Attach a per-campaign banner image. Same approved template can serve
    many use-cases (Diwali, Monsoon, Holiday…), each with its own banner.

    Flow:
      1. Validate file (≤5 MB, image MIME)
      2. Upload to Meta Media API → fresh media_id (lasts ~30 days)
      3. Store on the campaign — overrides template's default at send time
    """
    camp = db.get_campaign(campaign_id)
    if not camp:
        return jsonify({"ok": False, "error": "campaign not found"}), 404
    template = db.get_template_by_name(camp["template_name"])
    if not template or (template.get("header_type") or "").upper() != "IMAGE":
        return jsonify({"ok": False, "error": "this campaign's template has no image header"}), 400

    files = [f for f in request.files.getlist("image") if f.filename]
    if not files:
        return jsonify({"ok": False, "error": "no file provided"}), 400
    f = files[0]
    content = f.read()
    if not content:
        return jsonify({"ok": False, "error": "empty file"}), 400
    if len(content) > 5 * 1024 * 1024:
        return jsonify({"ok": False, "error": "image too large (max 5 MB)"}), 400
    mime = (f.mimetype or "").lower() or "image/jpeg"
    if not mime.startswith("image/"):
        return jsonify({"ok": False, "error": "file must be an image"}), 400

    api = WhatsAppAPI()
    if not api.is_configured():
        return jsonify({"ok": False, "error": "WhatsApp API not configured"}), 400

    media_id = api.upload_media(content, mime, filename=f"camp_{campaign_id}")
    if not media_id:
        return jsonify({"ok": False, "error": "Meta refused the upload (check size/format)"}), 502

    db.update_campaign_image(campaign_id, media_id, mime, len(content))
    return jsonify({"ok": True, "media_id": media_id, "size": len(content)})


@app.route("/campaigns/<int:campaign_id>/clear-image", methods=["POST"])
@login_required
def campaigns_clear_image(campaign_id):
    """Drop the per-campaign override; campaign reverts to template's default."""
    camp = db.get_campaign(campaign_id)
    if not camp:
        return jsonify({"ok": False, "error": "campaign not found"}), 404
    db.clear_campaign_image(campaign_id)
    return jsonify({"ok": True})


# ---------------- Review Follow-up ----------------
#
# Workflow:
#   1. RedPro emails a "RnR" (Ratings & Reviews) CSV to your Gmail —
#      OR you upload it directly.
#   2. App parses it, gets the set of phones that DID review.
#   3. App fetches recent bulk-send recipients from the messages table.
#   4. Non-reviewers = recent recipients minus reviewers.
#   5. Non-reviewers are loaded into the existing Send Preview page so you
#      pick a follow-up template (e.g. review_followup) and send / schedule.

@app.route("/reviews", methods=["GET"])
@login_required
def reviews():
    return render_template(
        "reviews.html",
        gmail_configured=gmail_fetch.is_configured(),
        gmail_address=db.get_setting("gmail_address", ""),
        gmail_sender_filter=db.get_setting("gmail_sender_filter", ""),
        gmail_subject_filter=db.get_setting("gmail_subject_filter", ""),
        default_lookback=db.get_setting("review_lookback_days", "14"),
    )


def _read_tabular_file(filename, content):
    """Return CSV-formatted bytes for either a CSV or Excel file."""
    ext = os.path.splitext(filename or "")[1].lower()
    if ext in (".xlsx", ".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        ws = wb.active
        out = io.StringIO()
        w = csv.writer(out)
        for row in ws.iter_rows(values_only=True):
            w.writerow([str(c) if c is not None else "" for c in row])
        return out.getvalue().encode("utf-8")
    return content


def _start_followup_batch(parsed, source_label):
    """Common: stash the parsed customers as pending → go to Send Preview."""
    passengers = parsed.get("passengers") or []
    total_rows = parsed.get("total_rows_seen", len(passengers))

    if not passengers:
        flash(
            f"No valid phone numbers found in '{source_label}'. "
            f"Check the file has a phone/mobile column.",
            "error",
        )
        return redirect(url_for("reviews"))

    # Respect opt-outs (same as the main Send flow)
    opted_out_phones = {r["phone"] for r in db.search_customers(opted_out=True, limit=50000)}
    before = len(passengers)
    passengers = [p for p in passengers if p["phone"] not in opted_out_phones]
    opt_out_filtered = before - len(passengers)

    pending = {
        "passengers": passengers,
        "stats": {
            "total_rows": total_rows,
            "cancelled":  parsed.get("cancelled_count", 0),
            "duplicates": parsed.get("duplicates_removed", 0),
            "invalid":    parsed.get("invalid_phones", 0),
            "opted_out":  opt_out_filtered,
            "ready":      len(passengers),
        },
        "filename": f"Review follow-up · {source_label}",
        "default_template": "review_followup",   # send_preview can pre-select
    }
    key = save_pending(pending)
    session["pending_key"] = key
    flash(
        f"✓ Loaded {len(passengers)} customers from '{source_label}'. "
        f"Pick your follow-up template below and send / schedule.",
        "success",
    )
    return redirect(url_for("send_preview"))


@app.route("/reviews/upload", methods=["POST"])
@login_required
def reviews_upload():
    files = [f for f in request.files.getlist("review_file") if f.filename]
    if not files:
        flash("Please pick a CSV or Excel file", "error")
        return redirect(url_for("reviews"))
    f = files[0]
    try:
        raw = f.read()
        content = _read_tabular_file(f.filename, raw)
        parsed = auto_parse(content, filename=f.filename)
    except Exception as e:
        flash(f"Failed to parse '{f.filename}': {e}", "error")
        return redirect(url_for("reviews"))
    return _start_followup_batch(parsed, f.filename)


@app.route("/reviews/fetch", methods=["POST"])
@login_required
def reviews_fetch():
    """Pull the next unread matching email from Gmail and start the follow-up batch."""
    if not gmail_fetch.is_configured():
        flash("Gmail not configured. Add address and App Password in Settings.", "error")
        return redirect(url_for("reviews"))
    try:
        emails = gmail_fetch.fetch_attachments(
            since_days=_safe_int(request.form.get("lookback_days", "14"), default=14, minimum=1),
            only_new=True,
            mark_seen=True,
        )
    except Exception as e:
        flash(f"Gmail fetch failed: {e}", "error")
        return redirect(url_for("reviews"))

    if not emails:
        flash("No new matching emails with attachments in Gmail.", "info")
        return redirect(url_for("reviews"))

    # Take the most recent matching email's first CSV/XLSX attachment
    latest = emails[-1]
    attach = next(
        (a for a in latest["attachments"]
         if (a.get("filename") or "").lower().endswith((".csv", ".xlsx", ".xls", ".txt"))),
        None,
    )
    if not attach:
        flash(f"Email '{latest.get('subject','')}' had no CSV/Excel attachment.", "error")
        return redirect(url_for("reviews"))

    try:
        content = _read_tabular_file(attach["filename"], attach["content_bytes"])
        parsed = auto_parse(content, filename=attach["filename"])
    except Exception as e:
        flash(f"Failed to parse '{attach['filename']}': {e}", "error")
        return redirect(url_for("reviews"))

    return _start_followup_batch(parsed, attach["filename"])


@app.route("/reviews/refine")
@login_required
def reviews_refine():
    """Filter & select non-reviewers before they're sent to send_preview."""
    key = session.get("pending_key")
    pending = load_pending(key)
    if not pending or not pending.get("passengers"):
        flash("No pending non-reviewers — upload a review CSV first.", "error")
        return redirect(url_for("reviews"))

    passengers = pending["passengers"]
    # Distinct values + counts for the filter sidebar
    from collections import Counter
    route_counts    = Counter((p.get("route") or "—") for p in passengers)
    platform_counts = Counter((p.get("platform") or "—") for p in passengers)
    template_counts = Counter((p.get("template_name") or "—") for p in passengers)

    # Date bounds for the date-range pickers (sent_at = "YYYY-MM-DD HH:MM:SS")
    dates = [p.get("sent_at") for p in passengers if p.get("sent_at")]
    min_date = min(dates)[:10] if dates else ""
    max_date = max(dates)[:10] if dates else ""

    return render_template(
        "reviews_refine.html",
        pending=pending,
        passengers=passengers,
        route_counts=sorted(route_counts.items(), key=lambda x: -x[1]),
        platform_counts=sorted(platform_counts.items(), key=lambda x: -x[1]),
        template_counts=sorted(template_counts.items(), key=lambda x: -x[1]),
        min_date=min_date,
        max_date=max_date,
    )


@app.route("/reviews/refine/continue", methods=["POST"])
@login_required
def reviews_refine_continue():
    """Take the user-refined passenger list and hand off to send_preview."""
    key = session.get("pending_key")
    pending = load_pending(key)
    if not pending:
        flash("Session expired — upload again.", "error")
        return redirect(url_for("reviews"))

    raw = request.form.get("passengers_json", "").strip()
    try:
        selected = json.loads(raw) if raw else []
    except Exception:
        selected = []
    selected = [p for p in selected if p.get("phone")]
    if not selected:
        flash("Pick at least one passenger before continuing.", "error")
        return redirect(url_for("reviews_refine"))

    # Reduce to the slim shape send_preview expects
    pending["passengers"] = [
        {
            "phone":    p["phone"],
            "name":     p.get("name") or "Customer",
            "route":    p.get("route") or "",
            "platform": p.get("platform") or "",
        }
        for p in selected
    ]
    pending["stats"]["ready"] = len(pending["passengers"])
    new_key = save_pending(pending)
    session["pending_key"] = new_key
    return redirect(url_for("send_preview"))


@app.route("/api/settings/gmail-test", methods=["POST"])
@login_required
def settings_gmail_test():
    ok, msg = gmail_fetch.test_connection()
    return jsonify({"success": ok, "message": msg})


@app.route("/api/campaigns/extract", methods=["POST"])
@login_required
def campaigns_extract():
    """AJAX: upload file(s) and return count of extracted phone numbers."""
    files = request.files.getlist("phone_file")
    files = [f for f in files if f.filename]
    all_phones = set()
    for f in files:
        content = f.read()
        ext = os.path.splitext(f.filename)[1].lower()
        if ext in (".xlsx", ".xls"):
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                text_parts = []
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        for cell in row:
                            if cell is not None:
                                text_parts.append(str(cell))
                content = " ".join(text_parts).encode("utf-8")
            except Exception:
                continue
        phones = extract_all_phones(content, f.filename)
        all_phones.update(phones)
    return jsonify({"count": len(all_phones), "phones_all": list(all_phones)})


# ---------------- Webhook (incoming WhatsApp messages) ----------------

def _dispatch_auto_reply(phone, msg, content):
    """Decide what (if anything) to auto-send back to a customer who just
    messaged us, then send it. Logged in the inbox like any outbound message.

    Logic, in order:
      1. Button/list reply → look up the FAQ text mapped to that button title
      2. Free-form text matching a configured keyword → send that FAQ text
      3. Anything else → send the welcome menu (up to 3 quick-reply buttons)

    Loop prevention: our own outbound messages never trigger the customer's
    webhook (Meta only emits 'messages' for inbound), so there's no echo
    risk. We still no-op when there's nothing useful to send.
    """
    api = WhatsAppAPI()
    if not api.is_configured():
        return

    msg_type = msg.get("type", "text")
    selected_label = None

    if msg_type == "interactive":
        inter = msg.get("interactive", {}) or {}
        if inter.get("type") == "button_reply":
            selected_label = inter.get("button_reply", {}).get("title", "")
        elif inter.get("type") == "list_reply":
            selected_label = inter.get("list_reply", {}).get("title", "")
    elif msg_type == "button":
        selected_label = msg.get("button", {}).get("text", "")

    reply_text = None
    if selected_label:
        # Button taps go straight to the configured response — no menu.
        match = db.find_auto_reply_by_text(selected_label)
        if match:
            reply_text = match["response_text"]
    elif msg_type == "text" and content:
        match = db.find_auto_reply_by_text(content)
        if match:
            reply_text = match["response_text"]

    if reply_text:
        ok, result = api.send_text(phone, reply_text)
        if ok:
            db.save_chat_message(
                phone, None, "out", reply_text,
                wa_message_id=result, message_type="text",
            )
        else:
            app.logger.error(f"Auto-reply text send failed to {phone}: {result}")
        return

    # No keyword/button match → send the welcome menu (if any buttons configured).
    if selected_label or msg_type not in ("text",):
        # Don't re-prompt with the menu in response to a button tap that
        # didn't resolve (would loop the customer back to the same menu)
        # or in response to non-text events like reactions/statuses.
        return

    menu = db.get_menu_button_replies()
    if not menu:
        return

    body_text = db.get_setting(
        "welcome_menu_body",
        "Welcome to Savan Travels 🚌\nHow can we help you today?",
    )
    footer = db.get_setting("welcome_menu_footer", "") or None

    # 1-3 menu items → buttons (single-tap UX).
    # 4-10 menu items → list message (modal with a CTA button, two-tap UX).
    if len(menu) <= 3:
        btn_payload = [
            {"id": f"ar_{b['id']}", "title": b["trigger_label"]} for b in menu
        ]
        ok, result = api.send_interactive_buttons(phone, body_text, btn_payload, footer)
        preview_tail = "[" + " | ".join(b["trigger_label"] for b in menu) + "]"
    else:
        rows = [
            {"id": f"ar_{b['id']}",
             "title": b["trigger_label"],
             "description": b.get("description") or ""}
            for b in menu
        ]
        button_label = db.get_setting("welcome_list_button_label", "Choose an option")
        ok, result = api.send_interactive_list(
            phone, body_text, rows,
            button_label=button_label, footer_text=footer,
        )
        preview_tail = "[List: " + " · ".join(b["trigger_label"] for b in menu) + "]"

    if ok:
        preview = body_text + "\n" + preview_tail
        db.save_chat_message(
            phone, None, "out", preview,
            wa_message_id=result, message_type="interactive",
        )
    else:
        app.logger.error(f"Welcome-menu send failed to {phone}: {result}")


@app.route("/webhook", methods=["GET"])
def webhook_verify():
    """Meta webhook verification handshake."""
    mode      = request.args.get("hub.mode")
    token     = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    stored    = db.get_setting("webhook_verify_token", "")
    if mode == "subscribe" and token == stored and stored:
        return challenge, 200
    return "Forbidden", 403


@app.route("/webhook", methods=["POST"])
def webhook_receive():
    """Receive incoming messages and status updates from Meta."""
    data = request.get_json(silent=True) or {}
    try:
        for entry in data.get("entry", []):
            for change in entry.get("changes", []):
                value = change.get("value", {})

                # ── Incoming messages ──────────────────────────────────────
                for msg in value.get("messages", []):
                    phone = msg.get("from", "")
                    wa_id = msg.get("id", "")
                    msg_type = msg.get("type", "text")

                    media_id = None
                    mime_type = None
                    filename = None
                    caption = ""
                    content = ""

                    if msg_type == "text":
                        content = msg.get("text", {}).get("body", "")
                    elif msg_type in ("image", "audio", "video", "document", "sticker"):
                        block = msg.get(msg_type, {}) or {}
                        media_id  = block.get("id")
                        mime_type = block.get("mime_type")
                        caption   = block.get("caption", "") or ""
                        if msg_type == "document":
                            filename = block.get("filename") or ""
                        content = caption or {
                            "image": "[Image]", "audio": "[Voice message]",
                            "video": "[Video]", "document": f"[Document: {filename or ''}]",
                            "sticker": "[Sticker]",
                        }[msg_type]
                    elif msg_type == "location":
                        loc = msg.get("location", {})
                        content = f"[Location: {loc.get('latitude')}, {loc.get('longitude')}]"
                    elif msg_type == "reaction":
                        # Defensive — sometimes Meta nests reaction info under
                        # different keys or omits it entirely on removal.
                        rx = msg.get("reaction") or {}
                        emoji = (rx.get("emoji") or "").strip()
                        content = f"{emoji} (reacted)" if emoji else "(reaction removed)"
                    elif msg_type == "button":
                        btn = msg.get("button", {})
                        content = f"🔘 {btn.get('text','') or btn.get('payload','') or 'Button reply'}"
                    elif msg_type == "interactive":
                        inter = msg.get("interactive", {})
                        if inter.get("type") == "button_reply":
                            content = f"🔘 {inter.get('button_reply',{}).get('title','Button reply')}"
                        elif inter.get("type") == "list_reply":
                            content = f"📋 {inter.get('list_reply',{}).get('title','List reply')}"
                        else:
                            content = f"[{msg_type}]"
                    else:
                        content = f"[{msg_type}]"

                    # Resolve customer name from contacts block
                    contacts = value.get("contacts", [])
                    name = contacts[0].get("profile", {}).get("name", "") if contacts else ""

                    # Download media now — Meta's URL expires in ~5 minutes.
                    # If R2 is configured, persist there (survives redeploys);
                    # otherwise fall back to the ephemeral local disk.
                    media_path = None
                    storage_kind = None
                    if media_id:
                        try:
                            api = WhatsAppAPI()
                            data, dl_mime = api.download_media(media_id)
                            if data:
                                if not mime_type:
                                    mime_type = dl_mime
                                ext = _ext_for_mime(mime_type)
                                safe_id = "".join(c for c in (wa_id or media_id) if c.isalnum() or c in "._-")
                                rel_name = f"{safe_id}{ext}"

                                if storage.is_configured():
                                    used = storage.upload_bytes(rel_name, data, mime_type)
                                    if used:
                                        media_path = rel_name
                                        storage_kind = used   # 'r2' or 'gridfs'
                                if not media_path:
                                    # Local fallback (also covers persistent-upload failure)
                                    full_path = os.path.join(MEDIA_DIR, rel_name)
                                    with open(full_path, "wb") as fh:
                                        fh.write(data)
                                    media_path = rel_name
                                    storage_kind = "local"
                        except Exception as e:
                            app.logger.error(f"Media download failed for {media_id}: {e}")

                    # Keep the raw Meta payload for anything non-text — so
                    # we never lose the original data even if our parser
                    # misses something. The inbox renders a fallback from
                    # this when content is a stale placeholder.
                    raw_payload = None
                    if msg_type != "text":
                        try:
                            raw_payload = json.dumps(msg, ensure_ascii=False)
                        except Exception:
                            raw_payload = None

                    db.save_chat_message(
                        phone, name or None, "in", content,
                        wa_message_id=wa_id, message_type=msg_type,
                        media_path=media_path, mime_type=mime_type, filename=filename,
                        storage_kind=storage_kind,
                        raw_payload=raw_payload,
                    )

                    # ── Auto-reply (if enabled in settings) ────────────────
                    # Behavior:
                    #   • inbound is a button/list reply → send the FAQ text
                    #     mapped to that label
                    #   • inbound is text matching a configured keyword → send
                    #     that FAQ text
                    #   • otherwise → send the welcome menu with up to 3
                    #     quick-reply buttons
                    if db.get_setting("auto_reply_enabled", "0") == "1":
                        try:
                            _dispatch_auto_reply(phone, msg, content)
                        except Exception as e:
                            app.logger.error(f"Auto-reply dispatch failed: {e}")

                # ── Status updates (sent → delivered → read) ───────────────
                for st in value.get("statuses", []):
                    db.update_chat_status(st.get("id", ""), st.get("status", ""))

    except Exception as e:
        app.logger.error(f"Webhook error: {e}")
    return jsonify({"status": "ok"}), 200


# ---------------- Inbox (two-way chat) ----------------

@app.route("/inbox")
@login_required
def inbox():
    conversations = db.list_conversations(50)
    raw_phone     = request.args.get("phone", "").strip()
    active_phone  = ""
    if raw_phone:
        from parsers import clean_phone as _cp
        cleaned = _cp(raw_phone)
        # Accept either a clean E.164 or pass through as-is if user typed
        # an already-formatted phone (e.g. 919..., 12 digits).
        active_phone = cleaned or raw_phone

    messages      = []
    customer_name = ""
    templates     = [dict(t) for t in db.get_templates()] if active_phone else []
    if active_phone:
        # Only the latest 50 of each on first paint — was 200 each which
        # produced 6 MB HTML on busy chats. The user can rarely scroll
        # 50 messages up anyway.
        chats = [dict(m) for m in db.get_conversation(active_phone, 50)]
        # Backfill reaction emojis from raw_payload when content is a stale
        # placeholder like "[reaction]" (e.g. from an older code path).
        for m in chats:
            if (m.get("message_type") == "reaction"
                and (not m.get("content") or m.get("content") == "[reaction]")
                and m.get("raw_payload")):
                try:
                    rp = json.loads(m["raw_payload"])
                    em = ((rp.get("reaction") or {}).get("emoji") or "").strip()
                    m["content"] = f"{em} (reacted)" if em else "(reaction removed)"
                except Exception:
                    pass
        bulk  = [dict(m) for m in db.get_phone_messages(active_phone, 50)]
        bodies = {t["name"]: t.get("body", "") for t in templates}
        for b in bulk:
            tname = b.get("template_name", "")
            rendered = render_template_body(
                bodies.get(tname, ""),
                name=b.get("customer_name"),
                route=b.get("route"),
                platform=b.get("platform"),
                template_name=tname,
            )
            content = rendered or f"📋 Sent template: {tname}"
            messages.append({
                "id": f"b{b.get('id')}",     # distinct id space from chats
                "direction": "out",
                "content": content,
                "timestamp": b.get("sent_at"),
                "status": b.get("status"),
                "customer_name": b.get("customer_name"),
                "message_type": "template",
                "template_name": tname,
                "is_bulk": True,
            })
        for c in chats:
            c["is_bulk"] = False
            messages.append(c)
        # Chronological merge
        messages.sort(key=lambda m: m.get("timestamp") or "")
        db.mark_read(active_phone)
        # Best-effort customer name (latest with a name wins)
        for m in reversed(messages):
            n = m.get("customer_name")
            if n:
                customer_name = n
                break
        if not customer_name:
            customer_name = active_phone

    # Last numeric chat id, used by JS for polling (bulk-only history → 0)
    last_chat_id = 0
    for m in messages:
        mid = m.get("id")
        if isinstance(mid, int) and mid > last_chat_id:
            last_chat_id = mid

    conversations = [dict(c) for c in conversations]
    return render_template("inbox.html",
                           conversations=conversations,
                           active_phone=active_phone,
                           messages=messages,
                           customer_name=customer_name,
                           templates=templates,
                           last_chat_id=last_chat_id)


@app.route("/api/inbox/send", methods=["POST"])
@login_required
def inbox_send():
    data    = request.get_json(silent=True) or {}
    phone   = data.get("phone", "").strip()
    text    = data.get("text", "").strip()
    if not phone or not text:
        return jsonify({"ok": False, "error": "phone and text required"}), 400

    api = WhatsAppAPI()
    if not api.is_configured():
        return jsonify({"ok": False, "error": "WhatsApp API not configured"}), 400

    ok, result = api.send_text(phone, text)
    if ok:
        msg_id = db.save_chat_message(
            phone, None, "out", text, wa_message_id=result, status="sent"
        )
        return jsonify({"ok": True, "id": msg_id})
    return jsonify({"ok": False, "error": result}), 400


@app.route("/api/inbox/send-template", methods=["POST"])
@login_required
def inbox_send_template():
    """Send an approved template to a phone from the Inbox.

    Body: { phone, template_name, params: [str, ...] (optional) }
    Used to initiate a chat when the 24-hour reply window is closed.
    """
    data = request.get_json(silent=True) or {}
    phone = (data.get("phone") or "").strip()
    tpl_name = (data.get("template_name") or "").strip()
    params_in = data.get("params") or []
    if not phone or not tpl_name:
        return jsonify({"ok": False, "error": "phone and template_name required"}), 400

    from parsers import clean_phone as _cp
    phone_clean = _cp(phone) or phone

    tpl = db.get_template_by_name(tpl_name)
    if not tpl:
        return jsonify({"ok": False, "error": "template not found"}), 404
    tpl = dict(tpl)
    if (tpl.get("status") or "").lower() not in ("approved", "active"):
        return jsonify({"ok": False, "error": f"template not approved (status: {tpl.get('status')})"}), 400

    # Build positional params, padding with "—" if user supplied fewer.
    var_count = int(tpl.get("variable_count") or 0)
    params = [str(p).strip() for p in params_in][:var_count]
    while len(params) < var_count:
        params.append("—")

    api = WhatsAppAPI()
    if not api.is_configured():
        return jsonify({"ok": False, "error": "WhatsApp API not configured"}), 400
    ok, result = api.send_template(
        phone_clean, tpl_name, tpl["language"], params,
        header_type=tpl.get("header_type"),
        header_example=tpl.get("header_example"),
        header_media_id=tpl.get("header_media_id"),
    )
    if not ok:
        return jsonify({"ok": False, "error": result}), 400

    # Render the body with the exact variable values used so the chat shows
    # what the customer actually received, not just the template name.
    body = tpl.get("body") or ""
    rendered = body
    for i, val in enumerate(params, start=1):
        rendered = rendered.replace("{{" + str(i) + "}}", str(val))
    content = rendered or f"📋 Sent template: {tpl_name}"

    msg_id = db.save_chat_message(
        phone_clean, None, "out", content,
        wa_message_id=result, message_type="template", status="sent",
    )
    return jsonify({"ok": True, "id": msg_id, "content": content})


@app.route("/api/inbox/messages")
@login_required
def inbox_poll():
    phone    = request.args.get("phone", "")
    after_id = _safe_int(request.args.get("after", 0), default=0, minimum=0)
    if not phone:
        return jsonify([])
    msgs = [dict(m) for m in db.get_new_messages(phone, after_id)]
    return jsonify(msgs)


@app.route("/api/inbox/unread")
@login_required
def inbox_unread():
    return jsonify({"count": db.get_unread_count()})


@app.route("/api/inbox/<phone>/delete", methods=["POST"])
@login_required
def inbox_delete_conversation(phone):
    """Hard-delete every chat message with this number. Cannot be undone."""
    from parsers import clean_phone as _cp
    phone = _cp(phone or "")
    if not phone:
        return jsonify({"ok": False, "error": "invalid phone"}), 400
    removed = db.delete_conversation(phone)
    return jsonify({"ok": True, "removed": removed})


# Meta's current actual tier ladder (visible in Business Manager UI).
# Older API responses may still report the legacy TIER_1K, so we map that
# to the closest current rung (2000) when picking the next rung.
_TIER_LADDER = [50, 250, 2_000, 10_000, 100_000]   # ascending


def _next_tier_size(current_limit):
    """Return the size of the next tier above `current_limit`, or None if at top."""
    if current_limit is None:
        return None
    for size in _TIER_LADDER:
        if size > current_limit:
            return size
    return None  # already past 100K → next is UNLIMITED (no numeric target)


def _label_for_size(size):
    """1000 -> '1K' ; 10000 -> '10K' ; 100000 -> '100K' ; None -> 'UNLIMITED'."""
    if size is None:
        return "UNLIMITED"
    if size >= 1_000_000:
        return f"{size // 1_000_000}M"
    if size >= 1_000:
        return f"{size // 1_000}K"
    return str(size)


@app.route("/api/dashboard/wa-limit")
@login_required
def dashboard_wa_limit():
    """Live tier + usage info for the WhatsApp limit card on the dashboard."""
    api = WhatsAppAPI()
    # ?fresh=1 from the Refresh button → skip cache. Normal page loads
    # use the 5-minute cache so the dashboard isn't gated on Meta's API.
    if request.args.get("fresh") == "1":
        info = api.get_phone_info() or {}
    else:
        info = api.get_phone_info_cached() or {}
    if info.get("error"):
        return jsonify({"ok": False, "error": info["error"]})

    used  = db.get_today_conversations()

    # The Meta Graph API's `messaging_limit_tier` field sometimes lags
    # behind the actual tier shown in Business Manager. Let the user pin
    # the real limit in Settings as a manual override. We pick the HIGHER
    # of (API value, override) so the dashboard auto-updates when Meta
    # actually upgrades the account, without the user touching settings.
    override_raw = (db.get_setting("whatsapp_tier_limit_override", "") or "").strip()
    override = None
    if override_raw:
        if override_raw.lower() in ("unlimited", "∞", "inf"):
            override = "unlimited"
        else:
            try:
                override = int(override_raw)
            except ValueError:
                override = None

    api_tier      = info.get("messaging_limit_tier") or ""
    api_limit     = info.get("tier_limit")          # int, or None when unlimited
    api_unlimited = (api_tier == "TIER_UNLIMITED")

    # "Unlimited" beats every numeric value
    if api_unlimited or override == "unlimited":
        limit          = None
        tier_label_str = "UNLIMITED"
        tier_source    = "api" if api_unlimited else "manual"
    else:
        manual_val = override if isinstance(override, int) and override > 0 else 0
        api_val    = api_limit or 0
        if manual_val > api_val:
            limit, tier_source = manual_val, "manual"
        elif api_val > 0:
            limit, tier_source = api_val, "api"
        else:
            limit, tier_source = None, "unknown"
        tier_label_str = _label_for_size(limit) if limit else "—"

    if limit:
        remaining = max(0, limit - used)
        pct = min(100, round(used * 100 / limit))
    else:
        remaining = None
        pct = 0

    # Upgrade-progress: 7-day unique customers vs next-rung target
    next_target = _next_tier_size(limit) if limit else None
    last7 = db.get_unique_conversations(days=7)
    next_remaining = None
    next_pct = 0
    if next_target:
        next_remaining = max(0, next_target - last7)
        next_pct = min(100, round(last7 * 100 / next_target))

    return jsonify({
        "ok":                True,
        "tier":              info.get("messaging_limit_tier") or "—",
        "tier_label":        tier_label_str,
        "tier_source":       tier_source,           # "api" or "manual"
        "limit":             limit,
        "used":              used,
        "remaining":         remaining,
        "percent":           pct,
        "quality":           info.get("quality_rating", "—"),
        "display_phone":     info.get("display_phone_number", ""),
        "verified_name":     info.get("verified_name", ""),
        # Upgrade path
        "next_tier_target":  next_target,
        "next_tier_label":   _label_for_size(next_target) if next_target else "UNLIMITED",
        "last7_unique":      last7,
        "next_tier_needed":  next_remaining,
        "next_tier_percent": next_pct,
    })


@app.route("/media/<int:chat_id>")
@login_required
def serve_media(chat_id):
    """Serve a customer-sent media file (image, audio, video, document)."""
    # Look up the chat record across whichever backend is active.
    chat = None
    if hasattr(db, "_db"):  # MongoDB backend exposes _db / _clean
        try:
            chat = db._clean(db._db().chats.find_one({"id": int(chat_id)}))
        except Exception:
            chat = None
    elif hasattr(db, "get_db"):  # SQLite fallback exposes a get_db() context manager
        try:
            with db.get_db() as conn:
                row = conn.execute(
                    "SELECT * FROM chats WHERE id=?", (int(chat_id),)
                ).fetchone()
                chat = dict(row) if row else None
        except Exception:
            chat = None

    if not chat or not chat.get("media_path"):
        abort(404)

    kind  = chat.get("storage_kind") or "local"
    name  = chat.get("filename") or chat["media_path"]
    mime  = chat.get("mime_type") or "application/octet-stream"

    # Cloudflare R2 → redirect to short-lived signed URL (browser fetches direct)
    if kind == "r2" and storage.is_r2_configured():
        url = storage.signed_url(chat["media_path"], filename=name)
        if url:
            return redirect(url)

    # MongoDB GridFS → proxy bytes through Flask (no public URL)
    if kind == "gridfs" and storage.is_gridfs_configured():
        data, gf_mime = storage.gridfs_read(chat["media_path"])
        if data:
            return send_file(
                io.BytesIO(data),
                mimetype=gf_mime or mime,
                as_attachment=False,
                download_name=name,
            )

    # Local file (also covers fallback if a remote backend failed to read)
    full = os.path.join(MEDIA_DIR, chat["media_path"])
    if not os.path.exists(full):
        abort(404)
    return send_file(
        full,
        mimetype=mime,
        as_attachment=False,
        download_name=name,
    )


# ---------------- Error handlers ----------------
@app.errorhandler(404)
def not_found(e):
    return render_template("error.html", code=404, message="Page not found"), 404

@app.errorhandler(500)
def server_error(e):
    return render_template("error.html", code=500, message="Something went wrong"), 500

@app.errorhandler(413)
def too_large(e):
    flash("File too large. Maximum 20MB allowed.", "error")
    return redirect(url_for("send"))


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print("=" * 55)
    print("🚌 Savan Travels WhatsApp Sender")
    print("=" * 55)
    print(f"📍 Open: http://localhost:{port}")
    print("👤 Login: admin / savan123")
    print("=" * 55)
    app.run(host="0.0.0.0", port=port, debug=False, use_reloader=False)
